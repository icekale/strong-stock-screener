from __future__ import annotations

import re
import warnings
from collections.abc import Mapping
from collections.abc import Sequence
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import date
from html.parser import HTMLParser
from io import BytesIO
from typing import Any, Generic, TypeVar

import httpx
from openpyxl import load_workbook

from app.models import (
    EtfHolderPosition,
    EtfSharePoint,
    MarginMarketPoint,
    StrongStockSourceStatus,
)


T = TypeVar("T")
_SSE_BASE_URL = "https://query.sse.com.cn"
_SZSE_BASE_URL = "https://www.szse.cn"
_HEADERS = {
    "Referer": "https://www.sse.com.cn/",
    "User-Agent": "Mozilla/5.0 (compatible; StrongStockWorkbench/1.0)",
}
_SINA_HOLDER_URL = (
    "https://stock.finance.sina.com.cn/fundInfo/api/openapi.php/"
    "CaihuiFundInfoService.getFundHolder"
)
_HOLDER_ENTITIES = frozenset(
    {
        "中央汇金投资有限责任公司",
        "中央汇金资产管理有限责任公司",
        "中国证券金融股份有限公司",
        "国新投资有限公司",
    }
)
_ISO_DATE_PATTERN = re.compile(r"(?<!\d)(\d{4}-\d{2}-\d{2})(?!\d)")


@dataclass(frozen=True)
class CapitalProviderResult(Generic[T]):
    rows: list[T]
    source_status: list[StrongStockSourceStatus]
    request_failures: int = 0
    available_trade_dates: tuple[str, ...] = field(default_factory=tuple)
    failed_symbols: tuple[str, ...] = field(default_factory=tuple)
    rejected_symbols: tuple[str, ...] = field(default_factory=tuple)
    empty_symbols: tuple[str, ...] = field(default_factory=tuple)


class OfficialCapitalDataProvider:
    def __init__(
        self,
        http_client: httpx.Client | None = None,
        timeout_seconds: float = 12,
    ) -> None:
        self.http_client = http_client or httpx.Client(timeout=timeout_seconds)

    def get_margin_rows(self, trade_date: str) -> CapitalProviderResult[MarginMarketPoint]:
        rows: list[MarginMarketPoint] = []
        statuses: list[StrongStockSourceStatus] = []
        compact_date = trade_date.replace("-", "")

        try:
            response = self.http_client.get(
                f"{_SSE_BASE_URL}/commonSoaQuery.do",
                params={
                    "isPagination": "true",
                    "sqlId": "RZRQ_HZ_INFO",
                    "beginDate": compact_date,
                    "endDate": compact_date,
                    "pageHelp.pageSize": 25,
                    "pageHelp.pageNo": 1,
                    "pageHelp.beginPage": 1,
                    "pageHelp.cacheSize": 1,
                    "pageHelp.endPage": 1,
                },
                headers=_HEADERS,
            )
            response.raise_for_status()
            sse_rows = parse_sse_margin_payload(response.json())
            rows.extend(sse_rows)
            statuses.append(_status("上交所两融", bool(sse_rows), "交易所汇总口径"))
        except Exception as exc:
            statuses.append(_failure_status("上交所两融", exc))

        try:
            response = self.http_client.get(
                f"{_SZSE_BASE_URL}/api/report/ShowReport/data",
                params={
                    "SHOWTYPE": "JSON",
                    "CATALOGID": "1837_xxpl",
                    "TABKEY": "tab1",
                    "txtDate": trade_date,
                },
                headers={**_HEADERS, "Referer": "https://www.szse.cn/"},
            )
            response.raise_for_status()
            szse_rows = parse_szse_margin_payload(response.json(), trade_date=trade_date)
            rows.extend(szse_rows)
            statuses.append(_status("深交所两融", bool(szse_rows), "交易所汇总口径"))
        except Exception as exc:
            statuses.append(_failure_status("深交所两融", exc))

        return CapitalProviderResult(rows=rows, source_status=statuses)


    def get_etf_share_rows(
        self,
        trade_date: str,
        symbols: Sequence[str],
    ) -> CapitalProviderResult[EtfSharePoint]:
        rows: list[EtfSharePoint] = []
        statuses: list[StrongStockSourceStatus] = []
        request_failures = 0
        available_trade_dates: set[str] = set()
        failed_symbols: set[str] = set()
        rejected_symbols: set[str] = set()
        empty_symbols: set[str] = set()
        sse_symbols = [symbol for symbol in symbols if symbol.endswith(".SH")]
        szse_symbols = [symbol for symbol in symbols if symbol.endswith(".SZ")]

        if sse_symbols:
            try:
                response = self.http_client.get(
                    f"{_SSE_BASE_URL}/commonQuery.do",
                    params={
                        "isPagination": "true",
                        "sqlId": "COMMON_SSE_ZQPZ_ETFZL_XXPL_ETFGM_SEARCH_L",
                        "STAT_DATE": trade_date,
                        "pageHelp.pageSize": 1000,
                        "pageHelp.pageNo": 1,
                        "pageHelp.beginPage": 1,
                        "pageHelp.cacheSize": 1,
                    },
                    headers=_HEADERS,
                )
                response.raise_for_status()
                sse_rows = parse_sse_etf_share_payload(response.json(), symbols=sse_symbols)
                rows.extend(sse_rows)
                available_trade_dates.update(row.trade_date for row in sse_rows)
                current_symbols = {
                    row.symbol for row in sse_rows if row.trade_date == trade_date
                }
                mismatched_symbols = {
                    row.symbol
                    for row in sse_rows
                    if row.trade_date != trade_date and row.symbol not in current_symbols
                }
                rejected_symbols.update(mismatched_symbols)
                empty_symbols.update(
                    set(sse_symbols) - current_symbols - mismatched_symbols
                )
                missing_count = len(set(sse_symbols) - current_symbols)
                returned_dates = sorted({row.trade_date for row in sse_rows})
                mismatched_dates = [date for date in returned_dates if date != trade_date]
                date_detail = (
                    f"；实际返回日期 {', '.join(returned_dates)}"
                    if mismatched_dates
                    else ""
                )
                statuses.append(
                    StrongStockSourceStatus(
                        source="上交所ETF份额",
                        status="success" if missing_count == 0 else "stale",
                        detail=(
                            f"当日有效 {len(current_symbols)}/{len(sse_symbols)} 只；"
                            f"缺失 {missing_count} 只；万份转换为份{date_detail}"
                        ),
                    )
                )
            except Exception as exc:
                request_failures += 1
                failed_symbols.update(sse_symbols)
                statuses.append(_failure_status("上交所ETF份额", exc))

        if szse_symbols:
            szse_rows: list[EtfSharePoint] = []
            szse_available_trade_dates: set[str] = set()
            failures: list[Exception] = []
            current = 0
            rejected = 0
            empty = 0
            for symbol in szse_symbols:
                try:
                    response = self.http_client.get(
                        f"{_SZSE_BASE_URL}/api/report/ShowReport/data",
                        params={
                            "SHOWTYPE": "JSON",
                            "CATALOGID": "1945",
                            "TABKEY": "tab1",
                            "txtQueryKeyAndJC": symbol.split(".", 1)[0],
                        },
                        headers={**_HEADERS, "Referer": "https://www.szse.cn/"},
                    )
                    response.raise_for_status()
                    payload = response.json()
                    section = _first_section(payload)
                    metadata = section.get("metadata") if isinstance(section, dict) else None
                    exchange_date = _szse_share_date(metadata)
                    if exchange_date is not None:
                        available_trade_dates.add(exchange_date)
                        szse_available_trade_dates.add(exchange_date)
                    if exchange_date is None or exchange_date != _date_text(trade_date):
                        rejected += 1
                        rejected_symbols.add(symbol)
                        continue
                    symbol_rows = parse_szse_etf_share_payload(
                        payload,
                        trade_date=trade_date,
                        symbols=[symbol],
                    )
                    if symbol_rows:
                        current += 1
                        szse_rows.extend(symbol_rows)
                    else:
                        empty += 1
                        empty_symbols.add(symbol)
                except Exception as exc:
                    failures.append(exc)
                    request_failures += 1
                    failed_symbols.add(symbol)
            rows.extend(szse_rows)
            actual_dates = sorted(szse_available_trade_dates)
            date_detail = f"；份额日期 {', '.join(actual_dates)}" if actual_dates else ""
            if failures and not szse_rows and not rejected and not empty:
                statuses.append(_failure_status("深交所ETF份额", failures[0]))
            elif current == len(szse_symbols):
                statuses.append(
                    _status(
                        "深交所ETF份额",
                        True,
                        f"当前快照归档；深市不补造历史{date_detail}",
                    )
                )
            else:
                statuses.append(
                    StrongStockSourceStatus(
                        source="深交所ETF份额",
                        status="stale",
                        detail=(
                            f"当日有效 {current}/{len(szse_symbols)} 只；"
                            f"{len(failures)} 只请求失败；{rejected + empty} 只日期或空数据拒绝；"
                            f"深市不补造历史{date_detail}"
                        ),
                    )
                )

        return CapitalProviderResult(
            rows=rows,
            source_status=statuses,
            request_failures=request_failures,
            available_trade_dates=tuple(sorted(available_trade_dates)),
            failed_symbols=tuple(symbol for symbol in symbols if symbol in failed_symbols),
            rejected_symbols=tuple(symbol for symbol in symbols if symbol in rejected_symbols),
            empty_symbols=tuple(symbol for symbol in symbols if symbol in empty_symbols),
        )

    def get_etf_share_history_rows(
        self,
        start_date: str,
        end_date: str,
        symbols: Sequence[str],
    ) -> CapitalProviderResult[EtfSharePoint]:
        szse_symbols = [symbol for symbol in symbols if symbol.endswith(".SZ")]
        if not szse_symbols:
            return CapitalProviderResult(rows=[], source_status=[])

        try:
            response = self.http_client.get(
                f"{_SZSE_BASE_URL}/api/report/ShowReport",
                params={
                    "SHOWTYPE": "xlsx",
                    "CATALOGID": "scsj_fund_jjgm",
                    "TABKEY": "tab1",
                    "txtStart": start_date,
                    "txtEnd": end_date,
                    "jjlb": "ETF",
                },
                headers={
                    **_HEADERS,
                    "Referer": "https://www.szse.cn/market/fund/volume/etf/index.html",
                },
            )
            response.raise_for_status()
            rows = parse_szse_daily_share_workbook(response.content, symbols=szse_symbols)
        except Exception as exc:
            return CapitalProviderResult(
                rows=[],
                source_status=[_failure_status("深交所ETF份额历史", exc)],
                request_failures=1,
                failed_symbols=tuple(szse_symbols),
            )

        rows = [row for row in rows if start_date <= row.trade_date <= end_date]
        covered_symbols = {row.symbol for row in rows}
        missing_symbols = set(szse_symbols) - covered_symbols
        available_dates = tuple(sorted({row.trade_date for row in rows}))
        status = "success" if rows and not missing_symbols else "stale"
        detail = (
            f"官方日频 {len(available_dates)} 个交易日；"
            f"覆盖 {len(covered_symbols)}/{len(szse_symbols)} 只"
        )
        return CapitalProviderResult(
            rows=rows,
            source_status=[
                StrongStockSourceStatus(
                    source="深交所ETF份额历史",
                    status=status,
                    detail=detail,
                )
            ],
            available_trade_dates=available_dates,
            empty_symbols=tuple(
                symbol for symbol in szse_symbols if symbol in missing_symbols
            ),
        )


class SinaEtfHolderProvider:
    def __init__(
        self,
        http_client: httpx.Client | None = None,
        timeout_seconds: float = 12,
        workers: int = 4,
    ) -> None:
        self.http_client = http_client or httpx.Client(timeout=timeout_seconds)
        self.workers = workers

    def get_holder_positions(
        self,
        symbols: Mapping[str, str],
    ) -> CapitalProviderResult[EtfHolderPosition]:
        positions: list[EtfHolderPosition] = []
        failures = 0
        with ThreadPoolExecutor(max_workers=min(self.workers, max(1, len(symbols)))) as executor:
            futures = {
                executor.submit(self._fetch_symbol, symbol, name): symbol
                for symbol, name in symbols.items()
            }
            for future in as_completed(futures):
                try:
                    positions.extend(future.result())
                except Exception:
                    failures += 1

        positions.sort(key=lambda item: (item.report_period, item.symbol, item.entity_name), reverse=True)
        latest_period = max((item.report_period for item in positions), default=None)
        if positions:
            status = "stale" if failures else "success"
            detail = f"最新报告期 {latest_period}，精确实体 {len(positions)} 条"
            if failures:
                detail += f"；{failures} 只ETF请求失败"
        else:
            status = "failed" if failures else "stale"
            detail = "未读取到核心ETF的精确实体持有人记录"
        return CapitalProviderResult(
            rows=positions,
            source_status=[
                StrongStockSourceStatus(source="新浪基金持有人", status=status, detail=detail)
            ],
        )

    def _fetch_symbol(self, symbol: str, name: str) -> list[EtfHolderPosition]:
        code = symbol.split(".", 1)[0]
        page_url = (
            "https://stock.finance.sina.com.cn/fundInfo/view/"
            f"FundInfo_JJCYR.php?symbol={code}"
        )
        response = self.http_client.get(page_url, headers=_HEADERS)
        response.raise_for_status()
        report_dates = parse_sina_report_dates(response.content.decode("gb18030", errors="replace"))
        if not report_dates:
            return []

        reports: dict[str, list[EtfHolderPosition]] = {}
        for report_period in report_dates[:2]:
            holder_response = self.http_client.get(
                _SINA_HOLDER_URL,
                params={"symbol": code, "date": report_period},
                headers={**_HEADERS, "Referer": page_url},
            )
            holder_response.raise_for_status()
            reports[report_period] = parse_sina_holder_payload(
                holder_response.json(),
                symbol=symbol,
                name=name,
                report_period=report_period,
            )

        latest_period = report_dates[0]
        previous_period = report_dates[1] if len(report_dates) > 1 else None
        previous_by_entity = {
            item.entity_name: item
            for item in reports.get(previous_period, [])
        }
        return [
            item.model_copy(
                update={
                    "change_shares": (
                        item.shares - previous_by_entity[item.entity_name].shares
                        if item.shares is not None
                        and previous_by_entity.get(item.entity_name) is not None
                        and previous_by_entity[item.entity_name].shares is not None
                        else None
                    )
                }
            )
            for item in reports.get(latest_period, [])
        ]


def parse_sse_margin_payload(payload: Any) -> list[MarginMarketPoint]:
    rows = _payload_rows(payload)
    output: list[MarginMarketPoint] = []
    for row in rows:
        trade_date = _date_text(row.get("opDate"))
        if not trade_date:
            continue
        output.append(
            MarginMarketPoint(
                trade_date=trade_date,
                market="SSE",
                financing_balance_cny=_number(row.get("rzye")),
                securities_lending_balance_cny=_number(row.get("rqylje")),
                margin_balance_cny=_number(row.get("rzrqjyzl")),
                financing_buy_cny=_number(row.get("rzmre")),
            )
        )
    return output


def parse_szse_margin_payload(payload: Any, *, trade_date: str) -> list[MarginMarketPoint]:
    section = _first_section(payload)
    rows = section.get("data") if isinstance(section, dict) else None
    if not isinstance(rows, list) or not rows or not isinstance(rows[0], dict):
        return []
    row = rows[0]
    return [
        MarginMarketPoint(
            trade_date=_date_text(trade_date) or trade_date,
            market="SZSE",
            financing_balance_cny=_hundred_million(row.get("jrrzye")),
            securities_lending_balance_cny=_hundred_million(row.get("jrrjye")),
            margin_balance_cny=_hundred_million(row.get("jrrzrjye")),
            financing_buy_cny=_hundred_million(row.get("jrrzmr")),
        )
    ]


def parse_sse_etf_share_payload(payload: Any, *, symbols: Sequence[str]) -> list[EtfSharePoint]:
    allowed = set(symbols)
    output: list[EtfSharePoint] = []
    for row in _payload_rows(payload):
        code = str(row.get("SEC_CODE") or "").strip()
        symbol = f"{code}.SH"
        total_shares = _ten_thousand(row.get("TOT_VOL"))
        trade_date = _date_text(row.get("STAT_DATE"))
        if symbol not in allowed or total_shares is None or total_shares <= 0 or not trade_date:
            continue
        output.append(
            EtfSharePoint(
                trade_date=trade_date,
                symbol=symbol,
                name=_plain_text(row.get("SEC_NAME")) or None,
                total_shares=total_shares,
            )
        )
    return output


def parse_szse_etf_share_payload(
    payload: Any,
    *,
    trade_date: str,
    symbols: Sequence[str],
) -> list[EtfSharePoint]:
    allowed = set(symbols)
    section = _first_section(payload)
    metadata = section.get("metadata") if isinstance(section, dict) else None
    exchange_date = _szse_share_date(metadata)
    if not exchange_date or exchange_date != _date_text(trade_date):
        return []
    rows = section.get("data") if isinstance(section, dict) else None
    if not isinstance(rows, list):
        return []
    output: list[EtfSharePoint] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        code = _plain_text(row.get("sys_key"))[:6]
        symbol = f"{code}.SZ"
        total_shares = _ten_thousand(row.get("dqgm"))
        if symbol not in allowed or total_shares is None or total_shares <= 0:
            continue
        output.append(
            EtfSharePoint(
                trade_date=exchange_date,
                symbol=symbol,
                name=_plain_text(row.get("kzjcurl")) or None,
                total_shares=total_shares,
                date_validation="szse_dqgm_v1",
            )
        )
    return output


def parse_szse_daily_share_workbook(
    payload: bytes,
    *,
    symbols: Sequence[str],
) -> list[EtfSharePoint]:
    allowed = set(symbols)
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("深交所日频份额响应不是有效 XLSX") from exc

    output: list[EtfSharePoint] = []
    found_columns = False
    try:
        for sheet in workbook.worksheets:
            sheet.reset_dimensions()
            columns: dict[str, int] | None = None
            for values in sheet.iter_rows(values_only=True):
                normalized = [_normalized_excel_header(value) for value in values]
                if columns is None:
                    try:
                        columns = {
                            "date": normalized.index("日期"),
                            "code": normalized.index("基金代码"),
                            "name": normalized.index("基金简称"),
                            "shares": next(
                                index
                                for index, value in enumerate(normalized)
                                if value in {"基金规模(份)", "基金份额"}
                            ),
                        }
                    except (ValueError, StopIteration):
                        continue
                    found_columns = True
                    continue

                trade_date = _workbook_date(values[columns["date"]])
                code = _fund_code(values[columns["code"]])
                symbol = f"{code}.SZ"
                total_shares = _number(values[columns["shares"]])
                if (
                    trade_date is None
                    or symbol not in allowed
                    or total_shares is None
                    or total_shares <= 0
                ):
                    continue
                output.append(
                    EtfSharePoint(
                        trade_date=trade_date,
                        symbol=symbol,
                        name=str(values[columns["name"]] or "").strip() or None,
                        total_shares=total_shares,
                        date_validation="szse_daily_v1",
                    )
                )
    finally:
        workbook.close()
    if not found_columns:
        raise ValueError("深交所日频份额工作簿缺少必要列")
    return sorted(output, key=lambda row: (row.trade_date, row.symbol))


def _szse_share_date(metadata: Any) -> str | None:
    if not isinstance(metadata, dict):
        return None
    columns = metadata.get("cols")
    title = columns.get("dqgm") if isinstance(columns, dict) else None
    values = _ISO_DATE_PATTERN.findall(str(title or ""))
    parsed_dates = [_date_text(value) for value in values]
    if not values or any(parsed is None for parsed in parsed_dates):
        return None
    dates = set(parsed_dates)
    return next(iter(dates)) if len(dates) == 1 else None


def parse_sina_report_dates(html: str) -> list[str]:
    parser = _SinaReportDateParser()
    parser.feed(html)
    return parser.dates


def parse_sina_holder_payload(
    payload: Any,
    *,
    symbol: str,
    name: str,
    report_period: str,
) -> list[EtfHolderPosition]:
    result = payload.get("result") if isinstance(payload, dict) else None
    rows = result.get("data") if isinstance(result, dict) else None
    if not isinstance(rows, list):
        return []
    positions: list[EtfHolderPosition] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        entity_name = str(row.get("cyrmc") or "").strip()
        if entity_name not in _HOLDER_ENTITIES:
            continue
        positions.append(
            EtfHolderPosition(
                symbol=symbol,
                name=name,
                report_period=report_period,
                entity_name=entity_name,
                shares=_number(row.get("cyfe")),
                holding_pct=_number(row.get("zfeb")),
                source="新浪财经基金持有人",
            )
        )
    return positions


def _payload_rows(payload: Any) -> list[dict[str, Any]]:
    if not isinstance(payload, dict):
        return []
    rows = payload.get("result")
    if not isinstance(rows, list):
        page_help = payload.get("pageHelp")
        rows = page_help.get("data") if isinstance(page_help, dict) else []
    return [row for row in rows if isinstance(row, dict)] if isinstance(rows, list) else []


def _first_section(payload: Any) -> dict[str, Any]:
    if isinstance(payload, list) and payload and isinstance(payload[0], dict):
        return payload[0]
    return {}


def _date_text(value: Any) -> str | None:
    text = str(value or "").strip()
    if len(text) == 8 and text.isdigit():
        text = f"{text[:4]}-{text[4:6]}-{text[6:]}"
    elif len(text) != 10 or text[4] != "-" or text[7] != "-":
        return None
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError:
        return None


def _workbook_date(value: Any) -> str | None:
    if isinstance(value, date):
        return value.date().isoformat() if hasattr(value, "date") else value.isoformat()
    return _date_text(value)


def _fund_code(value: Any) -> str:
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(6) if text.isdigit() else ""


def _normalized_excel_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).replace("（", "(").replace("）", ")")


def _number(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "--"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _hundred_million(value: Any) -> float | None:
    number = _number(value)
    return number * 100_000_000 if number is not None else None


def _ten_thousand(value: Any) -> float | None:
    number = _number(value)
    return number * 10_000 if number is not None else None


class _TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        self.parts.append(data)


class _SinaReportDateParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.inside_report_select = False
        self.dates: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if tag == "select" and attributes.get("id") == "tc_slt":
            self.inside_report_select = True
            return
        value = attributes.get("value")
        if tag == "option" and self.inside_report_select and _date_text(value):
            self.dates.append(str(value))

    def handle_endtag(self, tag: str) -> None:
        if tag == "select" and self.inside_report_select:
            self.inside_report_select = False


def _plain_text(value: Any) -> str:
    parser = _TextExtractor()
    parser.feed(str(value or ""))
    return "".join(parser.parts).strip()


def _status(source: str, has_rows: bool, detail: str) -> StrongStockSourceStatus:
    return StrongStockSourceStatus(
        source=source,
        status="success" if has_rows else "stale",
        detail=detail if has_rows else "交易所暂未返回当日数据",
    )


def _failure_status(source: str, exc: Exception) -> StrongStockSourceStatus:
    return StrongStockSourceStatus(
        source=source,
        status="failed",
        detail=f"请求失败: {exc.__class__.__name__}",
    )
